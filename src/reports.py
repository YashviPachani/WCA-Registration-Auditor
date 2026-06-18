# reports.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

RED    = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
YELLOW = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
GREEN  = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

def generate_report(missing_wca, missing_form, event_mismatches, duplicates, payment_df, output_path='report.xlsx'):
    wb = Workbook()

    def serialize(val):
        """Convert lists/sets to comma-separated strings for Excel."""
        if isinstance(val, (list, set)):
            return ', '.join(str(v) for v in val)
        return val

    def write_sheet(name, df, fill=None):
        ws = wb.create_sheet(name)
        if df.empty:
            ws.append(['No issues found'])
            return
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append([serialize(v) for v in row])
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill

    write_sheet('Missing in WCA',   missing_wca,             RED)
    write_sheet('Missing in Form',  missing_form,            RED)
    write_sheet('Event Mismatches', event_mismatches,        YELLOW)
    write_sheet('Duplicates Email', duplicates.get('email', pd.DataFrame()), YELLOW)
    write_sheet('Duplicates WCA ID',duplicates.get('wca_id', pd.DataFrame()), YELLOW)
    write_sheet('Payment Status',   payment_df)

    # Color payment sheet rows by status
    ps = wb['Payment Status']
    for row in ps.iter_rows(min_row=2):
        status_val = str(row[-1].value)
        if 'UNDERPAID' in status_val:
            for c in row: c.fill = RED
        elif 'OVERPAID' in status_val:
            for c in row: c.fill = YELLOW
        elif 'VERIFIED' in status_val:
            for c in row: c.fill = GREEN

    del wb['Sheet']
    wb.save(output_path)
    print(f'Report saved to {output_path}')

